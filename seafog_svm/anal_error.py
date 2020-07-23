from datetime import datetime
import os
from openpyxl import load_workbook

def anal_error_data(test_X, predicted_Y, test_Y):
    clock_cnt = [[0 for x in range(48)] for i in range(12)]

    for i, single_X in enumerate(test_X):
        timestamp = single_X[-1]
        dt = datetime.fromtimestamp(timestamp)

        if predicted_Y[i] != test_Y[i]:
            hour = dt.hour
            minute = dt.minute
            shift = 0
            if minute < 15: shift = 0
            elif minute > 45: shift = 2
            else: shift = 1
            clock_cnt[dt.month - 1][(2 * hour + shift) % 48] += 1
    return clock_cnt

def save_error(clock_cnt, title, tag):
    analysis_path = os.path.join('seafog_svm', 'data', 'output', title + '.xlsx')
    wb = load_workbook(analysis_path)
    wb.create_sheet(tag)
    sheet = wb.get_sheet_by_name(tag)
    for i in clock_cnt:
        sheet.append(i)
    wb.save(analysis_path)

def get_error_anal(test_X, predicted_Y, test_Y, title, tag):
    clock_cnt = anal_error_data(test_X, predicted_Y, test_Y)
    save_error(clock_cnt, title, tag)
